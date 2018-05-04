# The values in the array are calculated for each network reach using statistical models, calibrated from CHaMP data,
# which predict the proportion of habitat at or near each depth/velocity combination as a function of variables we know
# across the Upper Salmon network: bank-full width, gradient, and maybe discharge. (We don't have a discharge variable
# in the current shapefile.) The proportions from these regressions will be normalized so the whole array sums to 1,
# then multiplied by the surface area of the reach, so each entry in the array represents the actual surface area
# within the reach characterized by the given depth/velocity.

# ---------------------------------------------------------------------------------------------
# Code to start the boto3 session
# ---------------------------------------------------------------------------------------------

import boto3
import botocore
import pickle
from openpyxl import load_workbook

session = boto3.Session(profile_name='sfr')  # uses the credentials in ~/.aws/credentials [sfr] section
client = session.client('s3')

# ---------------------------------------------------------------------------------------------
# Code to list the keys to the dem_grid_results.csv files in the sfr-champdata bucket
# ---------------------------------------------------------------------------------------------
# paginator = client.get_paginator('list_objects_v2')
# page_iterator = paginator.paginate(
#     Bucket='sfr-champdata',
#     Prefix='QA/',
#     PaginationConfig={
#         'MaxItems': 1e12,
#         'PageSize': 1000
#     }
# )
# file_keys = []
# i = 0
# for page in page_iterator:
#     print("Loading page {0}".format(i))
#     file_keys += [item['Key'] for item in page['Contents'] if 'dem_grid_results.csv' in item['Key']]
#     i += 1

# Save result to file
# pickle.dump(file_keys, open('/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/bucket_keys_for_dem_grid_results.pickle', 'wb'))
# Load result from file
# file_keys = pickle.load(open('/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/bucket_keys_for_dem_grid_results.pickle', 'rb'))

# ---------------------------------------------------------------------------------------------
# Now download the grid results
# ---------------------------------------------------------------------------------------------

# upper_salmon_file_keys = [key for key in file_keys if 'YankeeFork' in key or 'Lemhi' in key]
#
# save_folder = '/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/dem_grid_results/'
# j = 0
# for key in upper_salmon_file_keys:
#     try:
#         year, watershed, sitename, visit, null1, null2, modelcode, null3 = key.split("/")[1:]
#         new_name = "{0}_{1}_{2}_{3}_{4}.csv".format(watershed, sitename, year, visit, modelcode)
#         print("Downloading file {0} of {1}: {2}".format(j+1, len(upper_salmon_file_keys), new_name))
#         client.download_file('sfr-champdata', key, save_folder + new_name)
#         j += 1
#     except botocore.exceptions.ClientError as e:
#         if e.response['Error']['Code'] == "404":
#             print("The object does not exist.")
#         else:
#             raise

# ---------------------------------------------------------------------------------------------
# Now figure out for which of these we have corresponding width (bankfull or average wetted), gradient, maybe stream order or mean discharge, and other data
# Ok, Carol is sending that...
# ---------------------------------------------------------------------------------------------

import pandas as pd
import os
import numpy as np

site_info_path = "/Users/Jason/Dropbox/SFR/CHaMP Database/20180207_metric_export_QATopo.csv"
hydro_results_folder = '/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/dem_grid_results/'
site_info = pd.DataFrame.from_csv(site_info_path)

# site_info['Site'], site_info['Area_Wet'], site_info['Area_Bf'], site_info['Grad']

upper_salmon_watersheds = ['Lemhi', 'Yankee Fork']

velocity_bin_edges = np.arange(-0.025, 1.0251, 0.05)  # bins of 0.05 m/s, except first is everything under 0.025 and last is everything over 0.975
depth_bin_edges = np.arange(-0.05, 2.1001, 0.1)  # bins of 0.1 m, except first is everything under 0.025 and last is everything over 0.975
velocity_bin_labels = np.arange(0, 1.0001, 0.05)
depth_bin_labels = np.arange(0, 2.0001, 0.1)
v_label_g, d_label_g = np.meshgrid(velocity_bin_labels, depth_bin_labels)

vel_depth_reg_data = {}

for index, row in site_info.iterrows():
    if row['Watershed'] in upper_salmon_watersheds:
        gradient = row['Grad']      # units? percent gradient (but expressed as proportion)
        width = row['BfWdth_Avg']   # units? m?
        prefix = "{0}_{1}".format(row['Watershed'].replace(" ", ""), row['Site'])
        hydro_files = [filename for filename in os.listdir(hydro_results_folder) if prefix in filename and "_S" in filename]
        if len(hydro_files) > 0:
            file_count = 0
            site_bincounts = np.zeros([len(velocity_bin_edges)-1, len(depth_bin_edges)-1])
            for hydro_file_name in hydro_files: # loop over all hydro files from the same site
                file_count += 1
                print("Processing site {3} of {4}, file {0} of {1}: {2}.".format(file_count, len(hydro_files), hydro_file_name, index, len(site_info)))
                site_data = pd.DataFrame.from_csv(hydro_results_folder + hydro_file_name)
                site_vd_data = np.array([site_data['Velocity.Magnitude'], site_data['Depth']]).T
                # Count the number of spatial grid cells within each depth/velocity bin for this file from this site
                site_bincounts += np.histogramdd(site_vd_data, (velocity_bin_edges, depth_bin_edges))[0]
            site_bincounts /= len(hydro_files)  # convert bincounts from sums to averages across all files from the same site
            # Count the number of spatial grid cells outside this depth/velocity bin for this site, for logistic regression misses/zeros
            site_otherbincounts = np.ones([len(velocity_bin_edges) - 1, len(depth_bin_edges) - 1]) * len(site_vd_data) - site_bincounts
            # Calculate proportions from above, in case we prefer to work with that
            site_binproportions = site_bincounts / len(site_vd_data)
            # Make sure another way of calculating the same thing matches
            assert (site_bincounts / (site_bincounts + site_otherbincounts) == site_binproportions).all()
            for i in range(len(velocity_bin_labels)):
                for j in range(len(depth_bin_labels)):
                    dv_count = site_bincounts[i, j]
                    otherdv_count = site_otherbincounts[i, j]
                    dv_proportion = site_binproportions[i, j]
                    key = "{0:.1f}_{1:.1f}".format(depth_bin_labels[i], velocity_bin_labels[j])
                    if key not in vel_depth_reg_data.keys():
                        vel_depth_reg_data[key] = []
                    vel_depth_reg_data[key].append([dv_proportion, dv_count, otherdv_count, gradient, width])
            break  # uncomment this break to only process one site for debugging

import pickle
pickle.dump(vel_depth_reg_data, open("/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/velocity_depth_regression_data.pickle", "wb"))

# NEXT STEP:

# Load the data from this pickle, which is a dict keyed by strings of "depth_velocity" with elements being arrays including the
# proportion of a site with that depth, the proportion of a site with that velocity (these proportions are the same, though,
# which they should be -- kind of an oversight but doesn't matter), gradient, and width. Then create a regression for each
# key to predict the proportion as a function of gradient and width.

import pickle
import numpy as np
import pandas as pd
import statsmodels.formula.api as sm
import statsmodels.tools as st
from SalNetIBM import betareg as br

vel_depth_reg_data = pickle.load(open("/Users/Jason/Dropbox/SFR/Projects/2018-02 NetworkHabitat/velocity_depth_regression_data.pickle", "rb"))

zmodels = {}  # logistic regression for probability that proportion is nonzero
bmodels = {}
zfits = {}  # ...
bfits = {}

for key, data in vel_depth_reg_data.items():
    print("Processing regressions for ", key)
    try:
        df = pd.DataFrame(data=data, index=np.arange(len(data)), columns=['proportion', 'hits', 'misses', 'gradient', 'width'])
        df['nonzero'] = df.apply(lambda row: 0 if row['proportion'] == 0 else 1, axis=1)
        df['intercept'] = df.apply(lambda row: 1, axis=1)
        df_nonzero = df.loc[df['nonzero'] == 1]
        if 0 < len(df_nonzero) < len(df):
            zmodels[key] = sm.Logit.from_formula('nonzero ~ gradient + width', data=df)
            zfits[key] = zmodels[key].fit()
            bmodels[key] = br.Beta.from_formula('proportion ~ gradient + width', data=df_nonzero) # does intercept automatically, despite documentation to the contrary
            bfits[key] = bmodels[key].fit()
        elif len(df_nonzero) == len(df):
            bmodels[key] = br.Beta.from_formula('proportion ~ gradient + width', data=df_nonzero) # does intercept automatically, despite documentation to the contrary
            bfits[key] = bmodels[key].fit()
    except st.sm_exceptions.PerfectSeparationError:
        print("PerfectSeparationError for key ", key)

def predict_proportions(dvkey, gradient, width):
    if dvkey in zfits.keys() and dvkey in bfits.keys():
        bprop = bmodels[dvkey].jrn_predict(bfits[dvkey].params[:-1], np.array([1, gradient, width]))
        temp_df = pd.DataFrame(data=[[gradient, width]], index=[0, 1], columns=['gradient', 'width'])
        zprop = zfits[dvkey].predict(temp_df)[0]
        return bprop * zprop
    elif dvkey in bfits.keys():
        bprop = bmodels[dvkey].jrn_predict(bfits[dvkey].params[:-1], np.array([1, gradient, width]))
        zprop = 1
        return bprop * zprop
    else:
        return 0

def normalized_proportions_for_reach(gradient, width):
    vals = np.array([(key, predict_proportions(key, gradient, width)) for key in vel_depth_reg_data.keys()], dtype=[('depth_velocity', np.unicode_, 7), ('proportion', np.float64)])
    vals['proportion'] = vals['proportion'] / vals['proportion'].sum()
    return dict(list(vals))

def habitat_areas_in_reach(gradient, width, wetted_area):
    proportions_dict = normalized_proportions_for_reach(gradient, width)
    return {key: value*wetted_area for key, value in proportions_dict.items()}

# Next parts of the plan...
# For any given fish size / temperature combination, calculate its NREI (with how much food???) for
# every depth/velocity combination and rank them by priority.


import matplotlib.pyplot as plt
import seaborn as sns
sns.set()
import random
plt.ioff()

######### This one just makes random plots to compare with lots of real sites
# for blah in range(100):
#     done = False
#     for index, row in site_info.iterrows():
#         if row['Watershed'] in upper_salmon_watersheds:
#             gradient = row['Grad']      # units? percent gradient (but expressed as proportion)
#             width = row['BfWdth_Avg']   # units? m?
#             prefix = "{0}_{1}".format(row['Watershed'].replace(" ", ""), row['Site'])
#             hydro_files = [filename for filename in os.listdir(hydro_results_folder) if prefix in filename and "_S" in filename]
#             if len(hydro_files) > 0:
#                 for hydro_file_name in hydro_files:  # loop over all hydro files from the same site
#                     site_data = pd.DataFrame.from_csv(hydro_results_folder + hydro_file_name)
#                     site_vd_data = np.array([site_data['Velocity.Magnitude'], site_data['Depth']]).T
#                     real_bincounts = np.histogramdd(site_vd_data, (velocity_bin_edges, depth_bin_edges))[0]
#                     if random.random() < 0.05:
#                         done = True
#                         break
#         if done:
#             break
#     real_proportions = real_bincounts / real_bincounts.sum()
#
#     binvalues = np.zeros([len(velocity_bin_labels), len(depth_bin_labels)])
#     nprop = normalized_proportions_for_reach(gradient, width)
#     for i in range(len(velocity_bin_labels)):
#         for j in range(len(depth_bin_labels)):
#             binvalues[i, j] = nprop['{0:.1f}_{1:.1f}'.format(depth_bin_labels[i], velocity_bin_labels[j])]
#     plotdf = pd.DataFrame(data=binvalues)
#     fig, (ax_model, ax_real) = plt.subplots(2, 1)
#     fig.set_size_inches(8.0, 16.0)
#     ax_real.set_xlabel('depth (m)')
#     ax_real.set_ylabel('velocity (m/s)')
#     ax_real.set_title('Random real site with same gradient/width')
#     ax_model.set_title('Modeled site with gradient {0:.2f} and width {1:.2f}'.format(gradient, width))
#     sns.heatmap(ax=ax_model, data=plotdf, xticklabels=depth_bin_labels, yticklabels=velocity_bin_labels, cmap='viridis', cbar_kws={'label': 'proportion'})
#     sns.heatmap(ax=ax_real, data=real_proportions, xticklabels=depth_bin_labels, yticklabels=velocity_bin_labels, cmap='viridis')
#     #plt.show()
#     plt.savefig('/Users/Jason/Desktop/tempfigs/grad{0:.1f}_width{1:.1f}.pdf'.format(gradient, width))

######### This one creates plots for a fixed gradient and variety of widths to see how extrapolation works
gradient = 1.0
for width in range(100):
    done = False
    binvalues = np.zeros([len(velocity_bin_labels), len(depth_bin_labels)])
    nprop = normalized_proportions_for_reach(gradient, width)
    for i in range(len(velocity_bin_labels)):
        for j in range(len(depth_bin_labels)):
            binvalues[i, j] = nprop['{0:.1f}_{1:.1f}'.format(depth_bin_labels[i], velocity_bin_labels[j])]
    plotdf = pd.DataFrame(data=binvalues)
    fig, ax_model = plt.subplots(1, 1)
    fig.set_size_inches(8.0, 8.0)
    ax_model.set_xlabel('depth (m)')
    ax_model.set_ylabel('velocity (m/s)')
    ax_model.set_title('Modeled site with gradient {0:.1f} % and width {1:.0f} m'.format(gradient, width))
    sns.heatmap(ax=ax_model, data=plotdf, xticklabels=depth_bin_labels, yticklabels=velocity_bin_labels, cmap='viridis', cbar_kws={'label': 'proportion'})
    plt.savefig('/Users/Jason/Desktop/widthfigs/grad{0:.1f}_width{1:04.0f}.pdf'.format(gradient, width))
plt.ion()


# ---------------------------------------------------------------------------------------------
# Generate temperature / fish size combinations for NREI rankings
# ---------------------------------------------------------------------------------------------

# Generate temperature/forkLength combinations as input files for BioenergeticHSC

import numpy as np
nrei_temperatures = np.arange(1, 21)
nrei_forkLengths = np.linspace(35, 600, 30)

import csv
with open('/Users/Jason/Dropbox/SFR/Projects/SalmonidNetworkIBM/resources/nrei_batch_specification.csv', 'w') as csvfile:
    writer = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Label', 'ForkLength', 'Mass', 'Temperature', 'Turbidity', 'DriftFile', 'Notes'])
    for temperature in nrei_temperatures:
        for forkLength in nrei_forkLengths:
            label = "t_{0:.0f}_fl_{1:.2f}".format(temperature, forkLength)
            mass = 10 ** (2.9 * np.log10(forkLength) - 4.7)
            writer.writerow([label, forkLength, mass, temperature, "", "", ""])

# Process an output file from BioenergeticHSC and rank depth/velocity combinations

import pandas as pd
import numpy as np
test_file = '/Users/Jason/Dropbox/SFR/Projects/SalmonidNetworkIBM/resources/nrei_batch_results_chena_drift/t_17_fl_151.90 (length 151.90 -- mass 42.31 -- temp 17.00).csv'

# accessing an element
df.loc[20.0].loc['20.0']  # first index row/velocity, second index column/depth as a string

# Create a list of ('depth_velocity', NREI) ordered by NREI descending (best NREIs first).
def habitat_preferences_from_file(habitat_pref_file):
    df = pd.read_csv(habitat_pref_file, index_col=0, dtype={'INDEX': np.float64})
    habitat_prefs = []
    for velocity in df.index:
        for depth in df.loc[velocity].index:
            nrei = df.loc[velocity].loc[str(depth)]
            if nrei > 0:
                label = "{0}_{1:.1f}".format(depth, velocity)
                habitat_prefs.append((label, nrei))
    return sorted(habitat_prefs, key=lambda x: -x[1])

habitat_preferences_from_file(test_file)

# ---------------------------------------------------------------------------------------------
# Process ALL the above output files into a single data structure so we can easily find a value
# for every fish at every timestep based on its growth, without loading from CSVs.
# ---------------------------------------------------------------------------------------------

BASE_DIRECTORY = os.path.join(os.path.expanduser("~"), 'Dropbox', 'SFR')
NREI_BATCH_FOLDER = os.path.join(BASE_DIRECTORY, 'Projects', 'SalmonidNetworkIBM', 'resources', 'nrei_batch_results')

habitat_preferences = {}
for root, dirs, filenames in os.walk(NREI_BATCH_FOLDER):
    for filename in filenames:
        label_parts = filename.split(' ')[0].split('_')
        temperature = int(label_parts[1])
        fork_length = float(label_parts[3])
        print("Temperature {0}, fork length {1:.2f}".format(temperature, fork_length))
        if temperature not in habitat_preferences.keys():
            habitat_preferences[temperature] = {}
        habitat_preferences[temperature][fork_length] = habitat_preferences_from_file(os.path.join(NREI_BATCH_FOLDER, filename))

# Now the code to access this dict for a given fish

def habitat_preferences_for_fish(temperature, fork_length):
    temperature_key = int(round(temperature))
    lengths_for_temperature = np.array(list(habitat_preferences[temperature_key].keys()))
    length_key = lengths_for_temperature[(np.abs(lengths_for_temperature - fork_length)).argmin()]
    return habitat_preferences[temperature_key][length_key]

test = habitat_preferences_for_fish(15, 25.3)

lengths_for_temperature = np.array(habitat_preferences[int(round(15))].keys())

# ---------------------------------------------------------------------------------------------
# DRIFT DATA
# ---------------------------------------------------------------------------------------------

# All files have SiteName, WatershedName, VisitID (which is sometimes shared among 2 replicates)

path_portion_sorted = "/Users/Jason/Dropbox/SFR/CHaMP Database/DriftInvertebrateSampleResultUpperSalmon.xlsx"  # PortionOfSampleSorted,
path_volume_sampled = "/Users/Jason/Dropbox/SFR/CHaMP Database/DriftInvertebrateSampleUpperSalmon.xlsx"  # MeasurementID (unique), VolumeSampled
path_bugs = "/Users/Jason/Dropbox/SFR/CHaMP Database/UpperSalmonTaxonBySizeClassAndCount.xlsx"  # MeasurementID, SizeClass, ObservedCount

wb_portion_sorted = load_workbook(path_portion_sorted, read_only=True)["DriftInvertebrateSampleResultUp"]["A1:AP1640"]
wb_volume_sampled = load_workbook(path_volume_sampled, read_only=True)["DriftInvertebrateSampleUpperSal"]["A1:AR401"]
wb_bugs = load_workbook(path_bugs, read_only=True)["UpperSalmonTaxonBySizeClassAndC"]["A1:AN12949"]

column_names_portion_sorted = [cell.value for cell in wb_portion_sorted[0]]
column_names_volume_sampled = [cell.value for cell in wb_volume_sampled[0]]
column_names_bugs = [cell.value for cell in wb_bugs[0]]

volumes_sampled = {}  # total N = 190, most add volumes from 1-2 nets per site, some up to 8
for row in wb_volume_sampled[1:]:
    values = [column.value for column in row]
    row_dict = dict(zip(column_names_volume_sampled, values))
    vid = row_dict['VisitID']
    vs = row_dict['VolumeSampled']
    if vs is not None:
        if vid in volumes_sampled.keys():
            volumes_sampled[vid] += vs
        else:
            volumes_sampled[vid] = vs

portions_sorted = {}  # total N = 202
for row in wb_portion_sorted[1:]:
    values = [column.value for column in row]
    row_dict = dict(zip(column_names_portion_sorted, values))
    vid = row_dict['VisitID']
    ps = row_dict['PortionOfSampleSorted']
    if ps is not None:
        portions_sorted[vid] = row_dict['PortionOfSampleSorted']

bugs = {}  # total N = 202
for row in wb_bugs[1:]:
    values = [column.value for column in row]
    row_dict = dict(zip(column_names_bugs, values))
    vid = row_dict['VisitID']
    if vid not in bugs.keys():
        bugs[vid] = []
    else:
        bugs[vid].append({key: row_dict.get(key) for key in ['LifeStage','Taxon','TaxonGroup', 'SizeClass', 'ObservedCount']})

# Narrow down to bug samples for which both volume and sorting data are available... this cuts out almost half
bugs = {key: value for key, value in bugs.items() if key in portions_sorted.keys() and key in volumes_sampled.keys()}

size_class_densities = {}
all_size_classes = []
for vid, vbugs in bugs.items():
    if vid not in size_class_densities.keys():
        size_class_densities[vid] = {}
    for bug in vbugs:
        sc = bug['SizeClass']
        if sc not in all_size_classes:
            all_size_classes.append(sc)
        if sc not in size_class_densities[vid].keys():
            size_class_densities[vid][sc] = bug['ObservedCount']
        else:
            size_class_densities[vid][sc] += bug['ObservedCount']
    for sc, count in size_class_densities[vid].items():
        size_class_densities[vid][sc] /= (volumes_sampled[vid] * portions_sorted[vid])
all_size_classes.sort(key=lambda scstring: int(scstring.split('-')[0]))

size_class_means = {sc: 0 for sc in all_size_classes}
size_class_maxes = {sc: 0 for sc in all_size_classes}
size_class_mins = {sc: 0 for sc in all_size_classes}

for vid, scdata in size_class_densities.items():
    for sc, density in scdata.items():
        size_class_means[sc] += density
        if density > size_class_maxes[sc]:
            size_class_maxes[sc] = density
        if density < size_class_mins[sc]:
            size_class_mins[sc] = density
for sc in all_size_classes:
    size_class_means[sc] /= len(size_class_densities)

import csv
with open('/Users/Jason/Dropbox/SFR/CHaMP Database/UpperSalmonDriftDensities.csv', 'w') as csvfile:
    writer = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Size class', 'Mean drift density (items/m3)', 'Min drift density', 'Max drift density'])
    for sc in all_size_classes:
        writer.writerow([sc + ' mm', size_class_means[sc], size_class_mins[sc], size_class_maxes[sc]])

# ---------------------------------------------------------------------------------------------
# Upper Salmon
# ---------------------------------------------------------------------------------------------

